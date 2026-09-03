import csv
from datetime import datetime
import os
from dotenv import load_dotenv
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook

from app.supabase_client import SupabaseStorage

load_dotenv()


class PaymentReportGenerator:
    """Build month-scoped payment records in PDF/Excel/CSV/DOC formats."""

    @staticmethod
    def _report_dir():
        report_dir = Path(__file__).resolve().parent.parent / "static" / "reports" / "payments"
        report_dir.mkdir(parents=True, exist_ok=True)
        return report_dir

    @staticmethod
    def _safe_text(value):
        if value is None:
            return ""
        return str(value).replace("\n", " ").replace("\r", " ")

    @staticmethod
    def _format_period(payment):
        if payment and getattr(payment, 'period', None):
            period = payment.period
        else:
            period = datetime.utcnow().strftime('%Y-%m')
        return period

    @staticmethod
    def _row_data(payment, transactions):
        rows = []
        total_amount = 0.0
        for tx in transactions or []:
            tx_row = tx if isinstance(tx, dict) else {
                'id': getattr(tx, 'id', ''),
                'customer': getattr(tx, 'customer', 'Unknown Customer'),
                'item': getattr(tx, 'item', 'Cafe Order'),
                'date': getattr(tx, 'date', ''),
                'amount': getattr(tx, 'amount', getattr(tx, 'order_price', 0.0)),
            }
            amount = float(tx_row.get('amount', 0.0) or 0.0)
            total_amount += amount
            rows.append({
                'order_id': tx_row.get('order_id', tx_row.get('id', '')),
                'customer': tx_row.get('customer', 'Unknown Customer'),
                'item': tx_row.get('item', 'Cafe Order'),
                'date': tx_row.get('date', ''),
                'amount': amount,
            })
        return rows, round(total_amount, 2)

    @staticmethod
    def _build_pdf(payment, rows, total_amount):
        period = PaymentReportGenerator._format_period(payment)
        timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "MSG Cafe Payment Record", ln=True, align='C')
        pdf.set_font("Arial", '', 10)
        pdf.cell(0, 10, f"Reporting Period: {period}", ln=True)
        pdf.cell(0, 10, f"Generated: {timestamp}", ln=True)
        pdf.cell(0, 10, f"Payment ID: {getattr(payment, 'id', 'N/A')}", ln=True)
        pdf.cell(0, 10, f"Payment Status: {getattr(payment, 'payment_status', 'unpaid')}", ln=True)
        pdf.ln(6)
        pdf.set_font("Arial", 'B', 10)
        pdf.set_fill_color(200, 220, 255)
        pdf.cell(25, 10, "Order ID", 1, 0, 'C', True)
        pdf.cell(50, 10, "Customer", 1, 0, 'C', True)
        pdf.cell(55, 10, "Item", 1, 0, 'C', True)
        pdf.cell(25, 10, "Date", 1, 0, 'C', True)
        pdf.cell(30, 10, "Amount", 1, 1, 'C', True)
        pdf.set_font("Arial", '', 9)
        for row in rows:
            pdf.cell(25, 8, PaymentReportGenerator._safe_text(row['order_id'])[:12], 1, 0, 'C')
            pdf.cell(50, 8, PaymentReportGenerator._safe_text(row['customer'])[:20], 1, 0, 'C')
            pdf.cell(55, 8, PaymentReportGenerator._safe_text(row['item'])[:18], 1, 0, 'C')
            pdf.cell(25, 8, PaymentReportGenerator._safe_text(row['date'])[:10], 1, 0, 'C')
            pdf.cell(30, 8, f"${row['amount']:.2f}", 1, 1, 'R')
        pdf.set_font("Arial", 'B', 11)
        pdf.cell(0, 10, f"Total Amount: ${total_amount:.2f}", ln=True, align='R')

        # 5. Upload generated PDF to Supabase Storage bucket.
        #file_name = f"reports/report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        #pdf_bytes = pdf.output(dest='S').encode('latin-1')

        #return SupabaseStorage.upload_bytes(file_name, pdf_bytes, "application/pdf")


        file_name = f"payment_{getattr(payment, 'id', 'record')}_{period.replace('-', '')}.pdf"
        output_path = PaymentReportGenerator._report_dir() / file_name
        pdf_out = pdf.output(dest='S')

        # fpdf returns a `str` or `bytes` depending on version
        if isinstance(pdf_out, str):
            pdf_bytes = pdf_out.encode('latin-1')
        else:
            pdf_bytes = pdf_out

        if os.getenv("FLASK_ENV") == "production":
            return SupabaseStorage.upload_bytes(file_name, pdf_bytes, "application/pdf")
        else:
            output_path.write_bytes(pdf_bytes)
            return str(output_path)
  

    @staticmethod
    def _build_csv(payment, rows, total_amount):
        period = PaymentReportGenerator._format_period(payment)
        file_name = f"payment_{getattr(payment, 'id', 'record')}_{period.replace('-', '')}.csv"
        output_path = PaymentReportGenerator._report_dir() / file_name
        with output_path.open('w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(['Payment ID', 'Reporting Period', 'Payment Status', 'Order ID', 'Customer', 'Item', 'Date', 'Amount'])
            for row in rows:
                writer.writerow([
                    getattr(payment, 'id', 'N/A'),
                    period,
                    getattr(payment, 'payment_status', 'unpaid'),
                    row['order_id'],
                    row['customer'],
                    row['item'],
                    row['date'],
                    f"{row['amount']:.2f}",
                ])
            writer.writerow(['', '', '', '', '', '', 'Total Amount', f"{total_amount:.2f}"])
        return str(output_path)

    @staticmethod
    def _build_xlsx(payment, rows, total_amount):
        period = PaymentReportGenerator._format_period(payment)
        file_name = f"payment_{getattr(payment, 'id', 'record')}_{period.replace('-', '')}.xlsx"
        output_path = PaymentReportGenerator._report_dir() / file_name

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Payment Record"
        headers = ['Payment ID', 'Reporting Period', 'Payment Status', 'Order ID', 'Customer', 'Item', 'Date', 'Amount']
        sheet.append(headers)
        for row in rows:
            sheet.append([
                getattr(payment, 'id', 'N/A'),
                period,
                getattr(payment, 'payment_status', 'unpaid'),
                row['order_id'],
                row['customer'],
                row['item'],
                row['date'],
                float(row['amount']),
            ])
        sheet.append(['', '', '', '', '', '', 'Total Amount', float(total_amount)])
        workbook.save(output_path)
        return str(output_path)

    @staticmethod
    def _build_doc(payment, rows, total_amount):
        period = PaymentReportGenerator._format_period(payment)
        file_name = f"payment_{getattr(payment, 'id', 'record')}_{period.replace('-', '')}.doc"
        output_path = PaymentReportGenerator._report_dir() / file_name
        lines = [
            "MSG Cafe Payment Record",
            f"Reporting Period: {period}",
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
            f"Payment ID: {getattr(payment, 'id', 'N/A')}",
            f"Payment Status: {getattr(payment, 'payment_status', 'unpaid')}",
            "",
            "Order ID | Customer | Item | Date | Amount",
        ]
        for row in rows:
            lines.append(f"{row['order_id']} | {row['customer']} | {row['item']} | {row['date']} | ${row['amount']:.2f}")
        lines.append(f"Total Amount: ${total_amount:.2f}")
        output_path.write_text("\n".join(lines), encoding='utf-8')
        return str(output_path)

    @staticmethod
    def generate_payment_document(payment, transactions, output_format='pdf'):
        rows, total_amount = PaymentReportGenerator._row_data(payment, transactions)
        format_name = (output_format or 'pdf').lower()
        if format_name == 'pdf':
            return PaymentReportGenerator._build_pdf(payment, rows, total_amount)
        if format_name in {'excel', 'xlsx'}:
            return PaymentReportGenerator._build_xlsx(payment, rows, total_amount)
        if format_name == 'csv':
            return PaymentReportGenerator._build_csv(payment, rows, total_amount)
        if format_name in {'doc', 'docx'}:
            return PaymentReportGenerator._build_doc(payment, rows, total_amount)
        raise ValueError(f"Unsupported payment export type: {output_format}")


class PDFGenerator:
    """
    Utility class for Process 4.2.
    Generates professional, auditable PDF reports for HR and Finance.
    """

    @staticmethod
    def generate_transaction_report(transactions, start_date, end_date):
        """
        Creates a PDF summary of all transactions within a specific range.
        Includes headers, detailed tables, and total revenue calculation.
        """
        pdf = FPDF()
        pdf.add_page()
        
        # 1. Header & Title
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "Automated Cafe Ordering System - Financial Report", ln=True, align='C')
        
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 10, f"Reporting Period: {start_date} to {end_date}", ln=True, align='C')
        pdf.cell(0, 10, f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", ln=True, align='C')
        pdf.ln(10)

        # 2. Table Headers
        # Defined widths for columns to ensure data alignment (Auditable layout)
        pdf.set_font("Arial", 'B', 10)
        pdf.set_fill_color(200, 220, 255) # Light blue background for header
        pdf.cell(30, 10, "ID", 1, 0, 'C', True)
        pdf.cell(60, 10, "Order Date", 1, 0, 'C', True)
        pdf.cell(60, 10, "Customer ID", 1, 0, 'C', True)
        pdf.cell(40, 10, "Amount (MWK)", 1, 1, 'C', True)

        # 3. Data Rows
        pdf.set_font("Arial", '', 10)
        total_revenue = 0
        
        for tx in transactions:
            pdf.cell(30, 8, str(tx.id), 1, 0, 'C')
            pdf.cell(60, 8, tx.created_at.strftime('%Y-%m-%d %H:%M'), 1, 0, 'C')
            pdf.cell(60, 8, str(tx.customer_id), 1, 0, 'C')
            pdf.cell(40, 8, f"{tx.order_price:,.2f}", 1, 1, 'R')
            total_revenue += tx.order_price

        # 4. Footer & Summary
        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(150, 10, "Total Gross Revenue:", 0, 0, 'R')
        pdf.cell(40, 10, f"{total_revenue:,.2f}", 1, 1, 'R')

        # 5. Upload generated PDF to Supabase Storage bucket.
        file_name = f"reports/report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf_bytes = pdf.output(dest='S').encode('latin-1')

        return SupabaseStorage.upload_bytes(file_name, pdf_bytes, "application/pdf")

    @staticmethod
    def generate_payment_export(payment, transactions, output_format='pdf'):
        return PaymentReportGenerator.generate_payment_document(payment, transactions, output_format)


ReportGenerator = PaymentReportGenerator
